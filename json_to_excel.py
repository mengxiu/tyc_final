import os, json
import openpyxl


def find_next_node(relationships: list, first_tyc_id: str):
    count = 0
    rela = []
    direction = 1
    for link in relationships[:]:
        if link['startNode'] == first_tyc_id:
            direction = 1
            rela.append(link['properties']['labels'])
            if count <= 0:
                nodes_list.append(link['endNode'])
            count += 1
            relationships.remove(link)
        if link['endNode'] == first_tyc_id:
            direction = -1
            rela.append(link['properties']['labels'])
            if count <= 0:
                nodes_list.append(link['startNode'])
            count += 1
            relationships.remove(link)
    direction_list.append(direction)
    rela_list.append(rela)


for file_fold in os.listdir('./result/json'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 80
    count = 0
    ws.cell(0 + 1 + count, 1).value = '根节点'
    ws.cell(0 + 1 + count, 2).value = '关系'
    ws.cell(0 + 1 + count, 3).value = '目标节点'
    count += 2
    for file_name in os.listdir('./result/json/{}'.format(file_fold)):
        with open('./result/json/{}/{}'.format(file_fold, file_name), 'r') as f:
            rela_dicts = json.load(f)
        if rela_dicts == {}:
            continue
        for k, _ in rela_dicts.items():
            if k == 'p_0':
                continue

            all_nodes_list = rela_dicts[k]['nodes']
            nodes_dict = {}
            nodes_list = []
            for i in all_nodes_list:
                nodes_dict.update({i['id']: i['properties']['name']})
                if i['properties']['name'].replace('）', ')').replace('（', '(') == file_fold or i['properties']['name']==file_fold :
                    nodes_list.append(i['id'])

            relationships = rela_dicts[k]['relationships'][:]
            rela_list = []
            direction_list = []
            while len(relationships) > 0:
                find_next_node(relationships, nodes_list[-1])
            nodes_list = list(map(lambda x: nodes_dict[x], nodes_list))
            print(nodes_list)

            for y in range(len(nodes_list) - 1):
                if direction_list[y] > 0:
                    ws.cell(y + 1 + count, 1).value = nodes_list[y]
                    ws.cell(y + 1 + count, 3).value = nodes_list[y + 1]
                else:
                    ws.cell(y + 1 + count, 1).value = nodes_list[y + 1]
                    ws.cell(y + 1 + count, 3).value = nodes_list[y]
                temp_str = ''
                for y_tmp in range(len(rela_list[y])):
                    for k_tmp in range(len(rela_list[y][y_tmp])):
                        if y_tmp > 0:
                            temp_str += ',' + rela_list[y][y_tmp][k_tmp]
                        else:
                            temp_str += rela_list[y][y_tmp][k_tmp]
                ws.cell(y + 1 + count, 2).value = temp_str
            count += len(nodes_list)
    if count <= 2:
        continue
    # del wb['Sheet']
    os.makedirs('./result/excel/', exist_ok=True)
    wb.save('./result/excel/{}.xlsx'.format(file_fold))
